# To works fully, use database filename as 'data.xlxs'


import openpyxl

wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

size_rows = sheet.max_row

data = {}

cont_udesc = 1
cont_externo = 1


print('[user]')
print()

for r in range(2,size_rows+1):
    print('usernumber=%d'%(r-1))
    print('usersitenumber=1')
    print('usericpcid=1')
    if sheet.cell(row=r, column=12).value == 'Remota':
        print('username=externo%d'%cont_externo)
        cont_externo += 1
    else:
        print('username=udesc%d'%cont_udesc)
        cont_udesc += 1
    tag = sheet.cell(row=r, column=9).value
    team_name = sheet.cell(row=r, column=2).value
    print('userfullname=[ %s ] %s'%(tag,team_name))
    print('usertype=team')
    password = sheet.cell(row=r, column=12).value
    password = str(password) if type(password)==str else str(int(password))
    print('userpassword=%s'%password)
    print('usermultilogin=t')
    print('userenabled=t')
    print()
        
lower_bound = ((cont_udesc+cont_externo)//10)*10
for r in range(lower_bound,100):
    id = str(r)
    print('usernumber=%d'%(r))
    print('usersitenumber=1')
    print('usericpcid=1')
    print('username=externo%s' % id)
    print('userfullname=[ AVULSO ] avulso%s' % id)
    print('usertype=team')
    print('userpassword=avulso%s' % id)
    print('usermultilogin=t')
    print('userenabled=t')
    print()